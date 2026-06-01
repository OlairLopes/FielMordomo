import datetime
import html
import io
from pathlib import Path

import pandas as pd


AZUL = "0B3A66"
AZUL_CLARO = "EAF2FB"
VERDE = "0F6E56"
VERMELHO = "C62828"
DOURADO = "D4AF37"


def _texto(valor) -> str:
    if valor is None or pd.isna(valor):
        return ""
    return str(valor)


def _moeda(valor) -> str:
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "R$ 0,00"


def _periodo_texto(periodo) -> str:
    inicio, fim = periodo
    return f"{inicio:%d/%m/%Y} a {fim:%d/%m/%Y}"


def _preparar_detalhes(df):
    detalhes = df.copy()
    if "data" in detalhes.columns:
        detalhes["data"] = pd.to_datetime(detalhes["data"], errors="coerce").dt.date
    remover = ["tipo_norm", "categoria_norm", "mes_ref"]
    return detalhes.drop(columns=[c for c in remover if c in detalhes.columns])


def gerar_excel_relatorio(
    df,
    resumo_categoria,
    resumo_subcategoria,
    igreja,
    periodo,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    detalhes = _preparar_detalhes(df)
    entradas = detalhes[detalhes["tipo"].str.upper() == "ENTRADA"]["valor"].sum()
    saidas = detalhes[detalhes["tipo"].str.upper() == "SAIDA"]["valor"].sum()
    saldo = entradas - saidas

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    titulo = Font(size=16, bold=True, color=AZUL)
    subtitulo = Font(size=11, italic=True, color="64748B")
    cabecalho = Font(bold=True, color="FFFFFF")
    fill_cabecalho = PatternFill("solid", fgColor=AZUL)
    fill_kpi = PatternFill("solid", fgColor=AZUL_CLARO)
    borda = Border(bottom=Side(style="thin", color="D7E3F0"))
    moeda_excel = 'R$ #,##0.00;[Red]-R$ #,##0.00'

    ws["A1"] = "Prestacao de Contas"
    ws["A1"].font = titulo
    ws["A2"] = _texto(igreja.get("nome", "Igreja"))
    ws["A2"].font = Font(size=13, bold=True, color=VERDE)
    ws["A3"] = f"Periodo: {_periodo_texto(periodo)}"
    ws["A3"].font = subtitulo
    ws["A4"] = f"Gerado em: {datetime.datetime.now():%d/%m/%Y %H:%M}"
    ws["A4"].font = subtitulo

    kpis = [
        ("Entradas", entradas),
        ("Saidas", saidas),
        ("Saldo", saldo),
        ("Lancamentos", len(detalhes)),
    ]
    for coluna, (rotulo, valor) in enumerate(kpis, start=1):
        cel_rotulo = ws.cell(6, coluna, rotulo)
        cel_valor = ws.cell(7, coluna, valor)
        cel_rotulo.font = Font(bold=True, color=AZUL)
        cel_rotulo.fill = fill_kpi
        cel_valor.fill = fill_kpi
        cel_valor.font = Font(size=12, bold=True, color=VERDE if rotulo != "Saidas" else VERMELHO)
        if rotulo != "Lancamentos":
            cel_valor.number_format = moeda_excel

    mensal = detalhes.copy()
    mensal["Mes"] = pd.to_datetime(mensal["data"], errors="coerce").dt.strftime("%Y-%m")
    mensal = (
        mensal.groupby(["Mes", "tipo"])["valor"]
        .sum()
        .unstack(fill_value=0.0)
        .rename(columns={"Entrada": "Entradas", "Saida": "Saidas"})
        .reset_index()
    )
    for coluna in ("Entradas", "Saidas"):
        if coluna not in mensal.columns:
            mensal[coluna] = 0.0
    mensal["Saldo"] = mensal["Entradas"] - mensal["Saidas"]

    inicio_tabela = 10
    ws.cell(inicio_tabela, 1, "Evolucao mensal").font = Font(bold=True, size=12, color=AZUL)
    for i, nome in enumerate(mensal.columns, start=1):
        celula = ws.cell(inicio_tabela + 1, i, nome)
        celula.font = cabecalho
        celula.fill = fill_cabecalho
    for r, row in enumerate(mensal.itertuples(index=False, name=None), start=inicio_tabela + 2):
        for c, valor in enumerate(row, start=1):
            celula = ws.cell(r, c, valor)
            celula.border = borda
            if c > 1:
                celula.number_format = moeda_excel

    if not mensal.empty:
        chart = BarChart()
        chart.title = "Entradas x Saidas por mes"
        chart.y_axis.title = "Valor (R$)"
        chart.x_axis.title = "Mes"
        dados = Reference(ws, min_col=2, max_col=3, min_row=inicio_tabela + 1, max_row=inicio_tabela + 1 + len(mensal))
        categorias = Reference(ws, min_col=1, min_row=inicio_tabela + 2, max_row=inicio_tabela + 1 + len(mensal))
        chart.add_data(dados, titles_from_data=True)
        chart.set_categories(categorias)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, "F10")

    def adicionar_aba_dataframe(nome, dataframe, moeda_colunas=()):
        aba = wb.create_sheet(nome)
        aba.sheet_view.showGridLines = False
        aba.freeze_panes = "A2"
        df_aba = dataframe.copy()
        for coluna in df_aba.columns:
            if pd.api.types.is_datetime64_any_dtype(df_aba[coluna]):
                df_aba[coluna] = df_aba[coluna].dt.date
        for c, coluna in enumerate(df_aba.columns, start=1):
            celula = aba.cell(1, c, str(coluna))
            celula.font = cabecalho
            celula.fill = fill_cabecalho
            celula.alignment = Alignment(horizontal="center")
        for r, row in enumerate(df_aba.itertuples(index=False, name=None), start=2):
            for c, valor in enumerate(row, start=1):
                celula = aba.cell(r, c, None if pd.isna(valor) else valor)
                celula.border = borda
                if df_aba.columns[c - 1] in moeda_colunas:
                    celula.number_format = moeda_excel
        if len(df_aba.columns):
            aba.auto_filter.ref = f"A1:{get_column_letter(len(df_aba.columns))}{max(1, len(df_aba) + 1)}"
        for c, coluna in enumerate(df_aba.columns, start=1):
            valores = [_texto(coluna)] + [_texto(v) for v in df_aba[coluna].head(200)]
            largura = min(42, max(12, max(len(v) for v in valores) + 2))
            aba.column_dimensions[get_column_letter(c)].width = largura
        return aba

    adicionar_aba_dataframe("Lancamentos", detalhes, moeda_colunas=("valor",))
    adicionar_aba_dataframe("Categorias", resumo_categoria, moeda_colunas=("Valor total",))
    adicionar_aba_dataframe("Despesas", resumo_subcategoria, moeda_colunas=("Valor total",))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A10"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def gerar_pdf_relatorio(
    df,
    resumo_categoria,
    resumo_subcategoria,
    igreja,
    periodo,
    logo=None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image, LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    detalhes = _preparar_detalhes(df)
    entradas = detalhes[detalhes["tipo"].str.upper() == "ENTRADA"]["valor"].sum()
    saidas = detalhes[detalhes["tipo"].str.upper() == "SAIDA"]["valor"].sum()
    saldo = entradas - saidas
    saida = io.BytesIO()

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloRelatorio", parent=estilos["Title"], textColor=colors.HexColor("#0B3A66"),
        fontSize=17, leading=20, alignment=TA_CENTER,
    )
    subtitulo = ParagraphStyle(
        "SubtituloRelatorio", parent=estilos["Normal"], textColor=colors.HexColor("#475569"),
        fontSize=9, leading=12, alignment=TA_CENTER,
    )
    secao = ParagraphStyle(
        "SecaoRelatorio", parent=estilos["Heading2"], textColor=colors.HexColor("#0B3A66"),
        fontSize=11, leading=14, spaceBefore=8, spaceAfter=5,
    )
    corpo = ParagraphStyle("CorpoRelatorio", parent=estilos["BodyText"], fontSize=7, leading=9)

    doc = SimpleDocTemplate(
        saida, pagesize=landscape(A4), leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.3 * cm,
        title="Prestacao de Contas", author="FielMordomo",
    )

    def rodape(canvas, documento):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(1.2 * cm, 0.7 * cm, "FielMordomo - Prestacao de contas")
        canvas.drawRightString(landscape(A4)[0] - 1.2 * cm, 0.7 * cm, f"Pagina {documento.page}")
        canvas.restoreState()

    historia = []
    if logo:
        try:
            dados, _ = logo
            imagem = Image(io.BytesIO(dados), width=2.6 * cm, height=1.5 * cm)
            imagem.hAlign = "CENTER"
            historia.append(imagem)
        except Exception:
            pass

    nome_igreja = html.escape(_texto(igreja.get("nome", "Igreja")))
    historia.extend([
        Paragraph("Prestacao de Contas", titulo),
        Paragraph(nome_igreja, titulo),
        Paragraph(
            f"Periodo: {_periodo_texto(periodo)} | Gerado em: {datetime.datetime.now():%d/%m/%Y %H:%M}",
            subtitulo,
        ),
        Spacer(1, 0.25 * cm),
    ])

    kpis = [
        ["Entradas", "Saidas", "Saldo", "Lancamentos"],
        [_moeda(entradas), _moeda(saidas), _moeda(saldo), str(len(detalhes))],
    ]
    tabela_kpi = Table(kpis, colWidths=[6.5 * cm] * 4)
    tabela_kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3A66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF2FB")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#0F6E56")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7E3F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    historia.extend([tabela_kpi, Spacer(1, 0.25 * cm)])

    def tabela_resumo(titulo_secao, dataframe, colunas):
        historia.append(Paragraph(titulo_secao, secao))
        linhas = [colunas]
        for _, row in dataframe.head(15).iterrows():
            linhas.append([Paragraph(html.escape(_texto(row[c])), corpo) for c in colunas])
        tabela = Table(linhas, repeatRows=1, colWidths=[11 * cm, 4 * cm, 5 * cm][:len(colunas)])
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3A66")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7E3F0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        historia.append(tabela)

    categorias_pdf = resumo_categoria.copy()
    categorias_pdf["Valor total"] = categorias_pdf["Valor total"].apply(_moeda)
    despesas_pdf = resumo_subcategoria.copy()
    despesas_pdf["Valor total"] = despesas_pdf["Valor total"].apply(_moeda)
    tabela_resumo("Resumo por categoria", categorias_pdf, ["Categoria", "Quantidade", "Valor total"])
    tabela_resumo("Despesas por subcategoria", despesas_pdf, ["Subcategoria", "Quantidade", "Valor total"])

    historia.append(Paragraph("Lancamentos detalhados", secao))
    linhas = [["Data", "Tipo", "Categoria", "Subcategoria", "Descricao", "Vinculado a", "Lote", "Valor"]]
    for _, row in detalhes.iterrows():
        data = row.get("data")
        data_txt = data.strftime("%d/%m/%Y") if hasattr(data, "strftime") else ""
        valores = [
            data_txt, row.get("tipo"), row.get("categoria"), row.get("subcategoria"),
            row.get("descricao"), row.get("nome_cadastro"), row.get("lote_id"),
            _moeda(row.get("valor")),
        ]
        linhas.append([Paragraph(html.escape(_texto(valor)), corpo) for valor in valores])

    tabela_detalhes = LongTable(
        linhas, repeatRows=1,
        colWidths=[1.8 * cm, 1.5 * cm, 2.2 * cm, 2.8 * cm, 5.2 * cm, 4.4 * cm, 3.2 * cm, 2.3 * cm],
    )
    tabela_detalhes.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3A66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#D7E3F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    historia.append(tabela_detalhes)
    doc.build(historia, onFirstPage=rodape, onLaterPages=rodape)
    return saida.getvalue()